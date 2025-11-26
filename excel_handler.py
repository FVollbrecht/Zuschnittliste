"""
Excel input/output handler for cutting optimization.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict
from pathlib import Path

from optimizer import Cut, Bar
from config import EXCEL_COLUMNS, INPUT_SHEET_NAME, OUTPUT_SHEET_NAME


class ExcelHandler:
    """Handles reading input and writing output Excel files."""
    
    @staticmethod
    def read_cuts_from_excel(file_path: str) -> List[Cut]:
        """
        Read cutting requirements from Excel file.
        
        Args:
            file_path: Path to input Excel file
            
        Returns:
            List of Cut objects
        """
        try:
            # Read the Excel file
            df = pd.read_excel(file_path, sheet_name=INPUT_SHEET_NAME, header=None)
            
            cuts = []
            
            # Process each row (skip header row 0)
            for idx, row in df.iterrows():
                if idx == 0:  # Skip header
                    continue
                
                try:
                    length = float(row[EXCEL_COLUMNS['length']])
                    quantity = int(row[EXCEL_COLUMNS['quantity']])
                    material_code = str(row[EXCEL_COLUMNS['material']]).strip()
                    material_name = str(row[EXCEL_COLUMNS['name']]).strip()
                    
                    # Validate data
                    if material_code and length > 0 and quantity > 0:
                        # Add each cut quantity times
                        for _ in range(quantity):
                            cuts.append(Cut(
                                length=length,
                                material_code=material_code,
                                material_name=material_name
                            ))
                except (ValueError, KeyError, IndexError):
                    # Skip invalid rows
                    continue
            
            return cuts
        
        except FileNotFoundError:
            raise FileNotFoundError(f"Input file not found: {file_path}")
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    @staticmethod
    def write_results_to_excel(results: Dict[str, Dict], output_path: str, bar_length: float):
        """
        Write optimization results to Excel file with formatting.
        
        Args:
            results: Dictionary with material codes as keys and optimization results as values
            output_path: Path for output Excel file
            bar_length: Standard bar length used
        """
        wb = Workbook()
        ws = wb.active
        ws.title = OUTPUT_SHEET_NAME
        
        # Styles
        title_font = Font(size=14, bold=True)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        material_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        
        current_row = 1
        
        # Main title
        ws.cell(row=current_row, column=1, value="Zuschnittoptimierung")
        ws.cell(row=current_row, column=1).font = title_font
        current_row += 2
        
        # Process each material
        for material_code, data in results.items():
            material_name = data['name']
            bars: List[Bar] = data['bars']
            
            if not bars:
                continue
            
            # Material header
            ws.cell(row=current_row, column=1, value="Material:")
            ws.cell(row=current_row, column=2, value=material_code)
            ws.cell(row=current_row, column=3, value=material_name)
            
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.font = header_font
                cell.fill = material_fill
            
            current_row += 2
            
            # Column headers
            headers = ["Stange", "Längen", "Gesamtlänge", "Rest", "Effizienz %"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
            
            # Bar details
            for bar in bars:
                # Bar number
                ws.cell(row=current_row, column=1, value=bar.bar_number)
                
                # Cuts (formatted as "length1 / length2 / ...")
                cuts_str = " / ".join(f"{cut:.1f}" for cut in bar.cuts)
                ws.cell(row=current_row, column=2, value=cuts_str)
                
                # Total used
                ws.cell(row=current_row, column=3, value=f"{bar.total_used:.1f}")
                
                # Waste
                ws.cell(row=current_row, column=4, value=f"{bar.waste:.1f}")
                
                # Efficiency
                ws.cell(row=current_row, column=5, value=f"{bar.efficiency:.1f}%")
                
                current_row += 1
            
            # Summary for this material
            total_bars = len(bars)
            total_cuts = sum(len(bar.cuts) for bar in bars)
            total_waste = sum(bar.waste for bar in bars)
            avg_efficiency = sum(bar.efficiency for bar in bars) / len(bars)
            
            current_row += 1
            ws.cell(row=current_row, column=1, value="Zusammenfassung:")
            ws.cell(row=current_row, column=1).font = header_font
            current_row += 1
            
            summary_data = [
                ("Anzahl Stangen:", total_bars),
                ("Anzahl Schnitte:", total_cuts),
                ("Gesamtverschnitt:", f"{total_waste:.1f} mm"),
                ("Durchschn. Effizienz:", f"{avg_efficiency:.1f}%")
            ]
            
            for label, value in summary_data:
                ws.cell(row=current_row, column=1, value=label)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1
            
            current_row += 2
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_path)
    
    @staticmethod
    def create_example_input(output_path: str = "example_input.xlsx"):
        """
        Create an example input Excel file with sample data.
        
        Args:
            output_path: Path for the example file
        """
        data = {
            'Länge (mm)': [2500, 1800, 1200, 900, 2400, 1500, 800, 2200, 1000, 1600],
            'Anzahl': [3, 5, 4, 6, 2, 4, 8, 3, 5, 4],
            'Material': ['ST37', 'ST37', 'ST37', 'ALU', 'ALU', 'ALU', 'ST52', 'ST52', 'ST52', 'ST52'],
            'Materialname': [
                'Stahl S235JR', 'Stahl S235JR', 'Stahl S235JR',
                'Aluminium 6060', 'Aluminium 6060', 'Aluminium 6060',
                'Stahl S355J2', 'Stahl S355J2', 'Stahl S355J2', 'Stahl S355J2'
            ]
        }
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=INPUT_SHEET_NAME, index=False)
